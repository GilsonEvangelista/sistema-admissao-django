from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.contrib.auth import authenticate, login
from django.contrib import messages
from django.shortcuts import render, redirect
from django.shortcuts import get_object_or_404
from django.shortcuts import redirect
from django.conf import settings
from .models import Candidato, Documento
import openpyxl


def login_view(request):

    if request.method == 'POST':

        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            return redirect('/dashboard/')
        else:
            return render(request, 'login.html', {'erro':'Usuário ou senha inválidos'})

    return render(request,'login.html')

@login_required
def lista_candidatos(request):

    busca = request.GET.get('busca')

    if busca:
        candidatos_lista = Candidato.objects.filter(nome__icontains=busca) | Candidato.objects.filter(ra__icontains=busca)
    else:
        candidatos_lista = Candidato.objects.all()

    paginator = Paginator(candidatos_lista, 10)  # 10 por página
    page = request.GET.get('page')
    candidatos = paginator.get_page(page)

    context = {
        'candidatos': candidatos
    }

    return render(request, 'lista_candidatos.html', context)

@login_required
def ficha_candidato(request, candidato_id):
    candidato = get_object_or_404(Candidato, id=candidato_id)
    context = {
        'candidato': candidato
    }

    return render (request, 'ficha_candidato.html', context)

@login_required
def dashboard(request):

    total_candidatos = Candidato.objects.count()
    total_documentos = Documento.objects.count()

    # Exemplo simples de status
    candidatos_com_docs = Documento.objects.values("candidato").distinct().count()
    candidatos_pendentes = total_candidatos - candidatos_com_docs

    if candidatos_pendentes > candidatos_com_docs:
        insight = "⚠️ Muitos candidatos ainda estão com documentos pendentes."
    else:
        insight = "✅ A maioria dos candidatos já enviou documentos."

    inscritos = Candidato.objects.filter(etapa="inscrito").count()
    documentos = Candidato.objects.filter(etapa="documentos").count()
    psicologia = Candidato.objects.filter(etapa="psicologia").count()
    entrevista = Candidato.objects.filter(etapa="entrevista").count()
    tacf = Candidato.objects.filter(etapa="tacf").count()
    medico = Candidato.objects.filter(etapa="medico").count()

    # Evolução (para gráfico)
    evolucao_labels = [
        "Inscrito", "Documentos", "Psicologia",
        "Entrevista", "TACF", "Médico"
    ]

    evolucao_values = [
        inscritos, documentos, psicologia,
        entrevista, tacf, medico
    ]
    # Contagem por tipo de documento
    documentos_por_tipo = {}

    for doc in Documento.objects.all():
        documentos_por_tipo[doc.tipo] = documentos_por_tipo.get(doc.tipo, 0) + 1

    context = {
        "total_candidatos": total_candidatos,
        "total_documentos": total_documentos,
        "candidatos_com_docs": candidatos_com_docs,
        "candidatos_pendentes": candidatos_pendentes,
        "doc_labels": list(documentos_por_tipo.keys()),
        "doc_values": list(documentos_por_tipo.values()),
        'insight': insight,
        'evolucao_labels': evolucao_labels,
        'evolucao_values': evolucao_values,

    }

    return render(request, "dashboard.html", context)

def limpar_cpf(cpf):
    import re
    cpf = re.sub(r'\D', '', str(cpf))  # remove tudo que não é número
    cpf = cpf.zfill(11)

    return cpf

@login_required
def importar_excel(request):

    if request.method == "POST":

        arquivo = request.FILES.get('arquivo')
        
        if not arquivo:
            return render(request, 'candidatos/importar.html', {
                'mensagem': 'Nenhum arquivo enviado.'
            })

        wb = openpyxl.load_workbook(arquivo)
        ws = wb.active

        headers = [str(cell.value).strip().upper() for cell in ws[1]]

        novos = 0
        atualizados = 0
        erros = 0
        erros_detalhados = []  # ✅ FORA do loop

        for row in ws.iter_rows(min_row=2, values_only=True):

            if not row:
                continue

            dados = dict(zip(headers, row))

            try:
                nome = str(dados.get("NOME-COMPLETO", "")).strip()
                cpf = limpar_cpf(dados.get("CPF", ""))
                ra = str(dados.get("RA", "")).strip()
                email = str(dados.get("EMAIL", "")).strip().lower() if dados.get("EMAIL") else ''
                etapa = str(dados.get("ETAPA", "inscrito")).strip().lower()

                # valida CPF
                if not cpf:
                    erros += 1
                    erros_detalhados.append(f"CPF inválido: {dados}")
                    continue

                obj, created = Candidato.objects.update_or_create(
                    cpf=cpf,
                    defaults={
                        'nome': nome,
                        'email': email,
                        'etapa': etapa,
                        'ra': ra,
                        'importado': True
                    }
                )  # ✅ FECHADO CORRETAMENTE

                if created:
                    novos += 1
                else:
                    atualizados += 1

            except Exception as e:
                erros += 1
                erros_detalhados.append(f"{dados} -> {str(e)}")
                continue

        return render(request, 'candidatos/importar.html', {
            'mensagem': f"✅ {novos} novos | 🔄 {atualizados} atualizados | ⚠️ {erros} erros",
            'erros_detalhados': erros_detalhados[:10]
        })

    return render(request, 'candidatos/importar.html', {
        'mensagem': ''
    })
    
@login_required
def limpar_candidatos(request):

    # 🔒 PROTEÇÃO: só funciona em modo DEBUG (desenvolvimento)
    if not settings.DEBUG:
        return redirect('dashboard')

    if request.method == "POST":
        total = Candidato.objects.count()
        Candidato.objects.all().delete()
        Documento.objects.all().delete()

        messages.success(request, f"{total} candidatos removidos com sucesso.")

    return redirect('dashboard')